#!/usr/bin/env python3
import argparse
import re
import sys
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))

from split_by_dong import save_dashboard_workbook


EXCLUDED_SHEETS = {"Summary", "홈페이지통계", "통합통계", "Data_Source"}
DERIVED_DONG_COLUMNS = {"세대 통합 메모"}


def normalize_header(value):
    return re.sub(r"\s+", "", str(value or ""))


def trim_dong_sheet_header(header):
    output = list(header)
    while output and (output[-1] is None or str(output[-1]).strip() in DERIVED_DONG_COLUMNS):
        output.pop()
    return output


def nonempty_row(row):
    return any(str(value or "").strip() for value in row)


def load_rows_from_dong_sheets(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    header = None
    data_rows = []

    for ws in wb.worksheets:
        if ws.title in EXCLUDED_SHEETS:
            continue
        if ws.max_row < 2:
            continue

        sheet_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        raw_header = trim_dong_sheet_header(sheet_header)
        if not raw_header:
            continue

        if header is None:
            header = raw_header
        elif [normalize_header(value) for value in header] != [normalize_header(value) for value in raw_header]:
            raise ValueError(f"{ws.title} 시트의 헤더가 다른 동 시트와 다릅니다.")

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_row = list(row[: len(raw_header)])
            if nonempty_row(raw_row):
                data_rows.append(raw_row)

    if header is None or not data_rows:
        raise ValueError("취합할 동별 시트를 찾지 못했습니다.")

    return header, data_rows


def parse_args():
    default_input = PROJECT_ROOT / "raw_data" / "발송현황_통계_동별분할_20260608.xlsx"
    parser = argparse.ArgumentParser(
        description="동별 시트의 수정 내용을 Data_Source로 취합하고 통계 시트를 다시 생성합니다."
    )
    parser.add_argument(
        "input",
        nargs="?",
        default=str(default_input),
        help="동별 시트가 포함된 엑셀 파일 경로",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="저장할 엑셀 파일 경로. 생략하면 입력 파일을 갱신합니다.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve() if args.output else input_path

    header, data_rows = load_rows_from_dong_sheets(input_path)
    save_dashboard_workbook(header, data_rows, input_path.name, output_path)

    print(f"Rebuilt dashboard workbook from dong sheets: {output_path}")
    print(f"Rows merged into Data_Source: {len(data_rows)}")


if __name__ == "__main__":
    main()
