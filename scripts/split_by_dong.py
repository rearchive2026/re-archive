import sys
import os
import re
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from math import ceil
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))
from import_verification_xlsx import load_xlsx_rows
from stats_source import (
    build_stats_column_map,
    cell_value,
    display_dong_label,
    final_submission,
    pick_column,
    row_with_excel_numeric_dong_ho,
    sort_dong_key,
)

PRIVACY_NEEDED_STATUS = "비대상"
DONE_STATUS = "완료"
SENT_STATUS = "발송"
VIEWED_STATUS = "열람"
RESERVED_STATUS = "예약"

CATEGORY_LABELS = [
    ("total", "통합 통계"),
    ("gyeongnam", "경남 아파트"),
    ("byeoksan", "벽산 아파트"),
]

OUTPUT_COLUMN_SPECS = [
    ("본번", "field", "bonbun"),
    ("동", "field", "dong"),
    ("호수", "field", "ho"),
    ("성명", "header", ["성명", "이름"]),
    ("공유여부", "field", "shared"),
    ("공유물건 동의상태", "header", ["공유물건 동의상태"]),
    ("다필지", "header", ["다필지"]),
    ("제출형태", "field", "submission"),
    ("최종동의방식", "final_submission", None),
    ("참여상태", "field", "participation"),
    ("발송 상태 메시지", "header", ["발송 상태 메시지"]),
    ("발송일시", "header", ["발송일시"]),
    ("열람일시", "header", ["열람일시"]),
    ("완료일시", "header", ["완료일시"]),
    ("메모", "field", "memo"),
]


def resolve_source_path(project_root):
    raw_data_path = project_root / "raw_data"
    dated_sources = []
    for path in raw_data_path.glob("발송현황_*.xlsx"):
        match = re.fullmatch(r"발송현황_(\d{8})\.xlsx", path.name)
        if match:
            dated_sources.append((match.group(1), path))

    if dated_sources:
        return max(dated_sources, key=lambda item: item[0])[1]
    return raw_data_path / "발송현황.xlsx"


def add_summary_charts(ws_summary, dong_header_row, dong_last_row):
    complex_chart = BarChart()
    complex_chart.title = "단지별 세대 완료율"
    complex_chart.y_axis.title = "완료율"
    complex_chart.y_axis.numFmt = "0.0%"
    complex_chart.y_axis.scaling.min = 0
    complex_chart.y_axis.scaling.max = 1
    complex_chart.x_axis.title = "구분"
    complex_chart.add_data(
        Reference(ws_summary, min_col=5, min_row=2, max_row=5),
        titles_from_data=True,
    )
    complex_chart.set_categories(Reference(ws_summary, min_col=1, min_row=3, max_row=5))
    complex_chart.width = 16
    complex_chart.height = 9
    ws_summary.add_chart(complex_chart, "Q1")

    dong_rate_chart = BarChart()
    dong_rate_chart.type = "bar"
    dong_rate_chart.title = "동별 세대 완료율"
    dong_rate_chart.x_axis.title = "완료율"
    dong_rate_chart.x_axis.numFmt = "0.0%"
    dong_rate_chart.x_axis.scaling.min = 0
    dong_rate_chart.x_axis.scaling.max = 1
    dong_rate_chart.y_axis.title = "동"
    dong_rate_chart.add_data(
        Reference(ws_summary, min_col=4, min_row=dong_header_row, max_row=dong_last_row),
        titles_from_data=True,
    )
    dong_rate_chart.set_categories(
        Reference(ws_summary, min_col=1, min_row=dong_header_row + 1, max_row=dong_last_row)
    )
    dong_rate_chart.width = 18
    dong_rate_chart.height = 14
    ws_summary.add_chart(dong_rate_chart, "Q16")

    shortage_chart = BarChart()
    shortage_chart.type = "bar"
    shortage_chart.title = "동별 50% 달성 부족분(서면)"
    shortage_chart.x_axis.title = "부족 세대"
    shortage_chart.y_axis.title = "동"
    shortage_chart.add_data(
        Reference(ws_summary, min_col=7, min_row=dong_header_row, max_row=dong_last_row),
        titles_from_data=True,
    )
    shortage_chart.set_categories(
        Reference(ws_summary, min_col=1, min_row=dong_header_row + 1, max_row=dong_last_row)
    )
    shortage_chart.width = 18
    shortage_chart.height = 14
    ws_summary.add_chart(shortage_chart, "Q38")


def normalize_row(row, width):
    return list(row) + [None] * max(0, width - len(row))


def apply_auto_filter(ws, header_row, start_col=1, end_col=None, end_row=None):
    end_col = end_col or ws.max_column
    end_row = end_row or ws.max_row
    if end_row < header_row or end_col < start_col:
        return
    ws.auto_filter.ref = (
        f"{get_column_letter(start_col)}{header_row}:"
        f"{get_column_letter(end_col)}{end_row}"
    )


def resolve_output_columns(header, data_rows, column_map):
    columns = []
    for label, source_type, source in OUTPUT_COLUMN_SPECS:
        if source_type == "field":
            index = column_map.get(source)
            columns.append((label, "index", index))
        elif source_type == "header":
            index = pick_column(header, data_rows, source)
            columns.append((label, "index", index))
        else:
            columns.append((label, source_type, None))
    return columns


def compact_output_row(row, output_columns, final_submission_value):
    values = []
    for _, source_type, index in output_columns:
        if source_type == "final_submission":
            values.append(final_submission_value)
        elif index is None or index >= len(row):
            values.append("")
        else:
            values.append(row[index])
    return values


def household_key(row, column_map):
    return (
        cell_value(row, column_map, "bonbun"),
        cell_value(row, column_map, "dong"),
        cell_value(row, column_map, "ho"),
    )


def classify_apt(bonbun):
    return "경남" if bonbun in {"525", "526"} else "벽산"


def category_key_for_bonbun(bonbun):
    return "gyeongnam" if bonbun in {"525", "526"} else "byeoksan"


def display_participation_label(status):
    labels = {
        PRIVACY_NEEDED_STATUS: "개인정보 동의 필요(비대상)",
        DONE_STATUS: "2차 전자 동의 완료",
        SENT_STATUS: "발송 후 미열람",
        VIEWED_STATUS: "열람 후 미완료",
        RESERVED_STATUS: "예약",
        "": "상태 없음",
    }
    return labels.get(status, status)


def create_homepage_stat_structure():
    return {
        "total_households": 0,
        "full_done_households": 0,
        "partial_done_households": 0,
        "not_done_households": 0,
        "privacy_needed_households": 0,
        "sent_households": 0,
        "viewed_households": 0,
        "reserved_households": 0,
        "shared_households": 0,
        "single_households": 0,
        "privacy_needed_members": 0,
        "sent_members": 0,
        "viewed_members": 0,
        "reserved_members": 0,
        "done_members": 0,
        "participation_details": defaultdict(int),
        "submission_details": defaultdict(int),
        "dong_stats": defaultdict(lambda: {"total": 0, "done": 0}),
    }


def finalize_homepage_stat(cat_data):
    count = cat_data["total_households"]
    dong_list = []
    for dong in sorted(cat_data["dong_stats"].keys(), key=sort_dong_key):
        total = cat_data["dong_stats"][dong]["total"]
        done = cat_data["dong_stats"][dong]["done"]
        dong_list.append(
            {
                "label": display_dong_label(dong),
                "total": total,
                "done": done,
                "rate": round(done / total * 100, 1) if total else 0,
            }
        )

    return {
        "count": count,
        "full_done_count": cat_data["full_done_households"],
        "full_done_rate": round(cat_data["full_done_households"] / count * 100, 1) if count else 0,
        "partial_done_count": cat_data["partial_done_households"],
        "remaining_household_count": count - cat_data["full_done_households"],
        "privacy_needed_household_count": cat_data["privacy_needed_households"],
        "privacy_needed_member_count": cat_data["privacy_needed_members"],
        "sent_household_count": cat_data["sent_households"],
        "sent_member_count": cat_data["sent_members"],
        "viewed_household_count": cat_data["viewed_households"],
        "viewed_member_count": cat_data["viewed_members"],
        "reserved_household_count": cat_data["reserved_households"],
        "reserved_member_count": cat_data["reserved_members"],
        "done_member_count": cat_data["done_members"],
        "shared_count": cat_data["shared_households"],
        "single_count": cat_data["single_households"],
        "household_stats": [
            {"label": "전원 완료 세대", "value": cat_data["full_done_households"]},
            {"label": "일부 완료 세대", "value": cat_data["partial_done_households"]},
            {"label": "미완료 세대", "value": cat_data["not_done_households"]},
        ],
        "owner_type_stats": [
            {"label": "단독 소유 세대", "value": cat_data["single_households"]},
            {"label": "공동 소유 세대", "value": cat_data["shared_households"]},
        ],
        "followup_details": [
            {
                "label": "1차 개인정보 동의 독려",
                "value": cat_data["privacy_needed_members"],
                "unit": "명",
                "sub": f"{cat_data['privacy_needed_households']}세대 포함",
            },
            {
                "label": "발송 후 미열람 리마인드",
                "value": cat_data["sent_members"],
                "unit": "명",
                "sub": f"{cat_data['sent_households']}세대 포함",
            },
            {
                "label": "열람 후 미완료 안내",
                "value": cat_data["viewed_members"],
                "unit": "명",
                "sub": f"{cat_data['viewed_households']}세대 포함",
            },
            {
                "label": "일부 완료 공유세대 추가 확인",
                "value": cat_data["partial_done_households"],
                "unit": "세대",
                "sub": "전원 완료 전환 가능 세대",
            },
            {
                "label": "예약 상태 확인",
                "value": cat_data["reserved_members"],
                "unit": "명",
                "sub": f"{cat_data['reserved_households']}세대 포함",
            },
        ],
        "participation_details": sorted(
            [{"label": label, "value": value} for label, value in cat_data["participation_details"].items()],
            key=lambda item: item["value"],
            reverse=True,
        ),
        "submission_details": sorted(
            [{"label": label, "value": value} for label, value in cat_data["submission_details"].items()],
            key=lambda item: item["value"],
            reverse=True,
        ),
        "dong_stats": dong_list,
    }


def build_homepage_stats(padded_rows, column_map, final_submissions):
    household_members = defaultdict(list)
    for index, row in enumerate(padded_rows):
        household_members[household_key(row, column_map)].append((index, row))

    stats = {
        "total": create_homepage_stat_structure(),
        "gyeongnam": create_homepage_stat_structure(),
        "byeoksan": create_homepage_stat_structure(),
    }

    for (bonbun, dong, ho), members in household_members.items():
        category_key = category_key_for_bonbun(bonbun)
        member_statuses = [cell_value(row, column_map, "participation") for _, row in members]
        done_count = sum(status == DONE_STATUS for status in member_statuses)
        is_shared_household = any("공유 O" in cell_value(row, column_map, "shared") for _, row in members) or len(members) > 1
        has_privacy_needed = any(status == PRIVACY_NEEDED_STATUS for status in member_statuses)
        has_sent = any(status == SENT_STATUS for status in member_statuses)
        has_viewed = any(status == VIEWED_STATUS for status in member_statuses)
        has_reserved = any(status == RESERVED_STATUS for status in member_statuses)

        household_status = "미동의"
        if done_count == len(members):
            household_status = "전원완료"
        elif done_count > 0:
            household_status = "일부완료"

        for key in ["total", category_key]:
            current = stats[key]
            current["total_households"] += 1
            if is_shared_household:
                current["shared_households"] += 1
            else:
                current["single_households"] += 1

            if household_status == "전원완료":
                current["full_done_households"] += 1
            elif household_status == "일부완료":
                current["partial_done_households"] += 1
            else:
                current["not_done_households"] += 1

            if has_privacy_needed:
                current["privacy_needed_households"] += 1
            if has_sent:
                current["sent_households"] += 1
            if has_viewed:
                current["viewed_households"] += 1
            if has_reserved:
                current["reserved_households"] += 1

            dong_key = dong or "기타"
            current["dong_stats"][dong_key]["total"] += 1
            if household_status == "전원완료":
                current["dong_stats"][dong_key]["done"] += 1

            for index, row in members:
                participation = cell_value(row, column_map, "participation")
                current["participation_details"][display_participation_label(participation)] += 1
                current["submission_details"][final_submissions[index]] += 1
                if participation == PRIVACY_NEEDED_STATUS:
                    current["privacy_needed_members"] += 1
                elif participation == SENT_STATUS:
                    current["sent_members"] += 1
                elif participation == VIEWED_STATUS:
                    current["viewed_members"] += 1
                elif participation == RESERVED_STATUS:
                    current["reserved_members"] += 1
                elif participation == DONE_STATUS:
                    current["done_members"] += 1

    return {key: finalize_homepage_stat(value) for key, value in stats.items()}


def build_dashboard_rows(data_rows, header_width, column_map):
    padded_rows = [
        row_with_excel_numeric_dong_ho(
            normalize_row(row, header_width),
            column_map,
        )
        for row in data_rows
    ]
    household_members = defaultdict(list)
    for index, row in enumerate(padded_rows):
        household_members[household_key(row, column_map)].append(index)

    final_submissions = [""] * len(padded_rows)
    household_summaries = {}

    for key, indexes in household_members.items():
        statuses = [cell_value(padded_rows[index], column_map, "participation") for index in indexes]
        submissions = [
            final_submission(padded_rows[index], column_map)
            for index in indexes
        ]
        for index, submission in zip(indexes, submissions):
            final_submissions[index] = submission
        done_count = sum(status == "완료" for status in statuses)
        written_count = sum(submission == "서면" for submission in submissions)
        size = len(indexes)
        is_full_done = int(size > 0 and done_count == size)
        is_written_done = int(is_full_done and written_count == size)
        is_partial = int(0 < done_count < size)
        bonbun, dong, ho = key
        household_summaries[key] = {
            "bonbun": bonbun,
            "dong": dong or "기타",
            "ho": ho,
            "apt": classify_apt(bonbun),
            "size": size,
            "done_count": done_count,
            "written_count": written_count,
            "is_full_done": is_full_done,
            "is_written_done": is_written_done,
            "is_partial": is_partial,
        }

    return padded_rows, final_submissions, household_summaries


def summarize_complex(padded_rows, household_summaries, column_map, apt=None):
    if apt:
        member_rows = [row for row in padded_rows if classify_apt(cell_value(row, column_map, "bonbun")) == apt]
        households = [summary for summary in household_summaries.values() if summary["apt"] == apt]
    else:
        member_rows = padded_rows
        households = list(household_summaries.values())

    member_count = len(member_rows)
    done_members = sum(cell_value(row, column_map, "participation") == "완료" for row in member_rows)
    household_count = len(households)
    full_done_count = sum(summary["is_full_done"] for summary in households)
    written_done_count = sum(summary["is_written_done"] for summary in households)
    partial_count = sum(summary["is_partial"] for summary in households)

    return {
        "member_count": member_count,
        "member_done_rate": done_members / member_count if member_count else 0,
        "household_count": household_count,
        "household_done_rate": full_done_count / household_count if household_count else 0,
        "written_done_count": written_done_count,
        "partial_count": partial_count,
    }


def summarize_dongs(household_summaries):
    grouped = defaultdict(list)
    for summary in household_summaries.values():
        grouped[summary["dong"]].append(summary)

    result = []
    for dong in sorted(grouped, key=sort_dong_key):
        households = grouped[dong]
        total = len(households)
        full_done = sum(summary["is_full_done"] for summary in households)
        written_done = sum(summary["is_written_done"] for summary in households)
        partial = sum(summary["is_partial"] for summary in households)
        shortage = max(0, ceil(total / 2) - written_done)
        result.append(
            {
                "dong": dong,
                "total": total,
                "full_done": full_done,
                "rate": full_done / total if total else 0,
                "written_done": written_done,
                "partial": partial,
                "shortage": shortage,
            }
        )
    return result


def style_range(ws, row, start_col, end_col, fill=None, font=None, border=None, alignment=None):
    for col in range(start_col, end_col + 1):
        target = ws.cell(row=row, column=col)
        if fill:
            target.fill = fill
        if font:
            target.font = font
        if border:
            target.border = border
        if alignment:
            target.alignment = alignment


def append_homepage_table(ws, title, headers, rows, start_row, styles):
    ws.cell(start_row, 1, title)
    style_range(ws, start_row, 1, len(headers), font=styles["section_font"])
    start_row += 1

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.border = styles["border"]
        cell.alignment = styles["center_align"]
    start_row += 1

    for row_values in rows:
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(start_row, col, value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        start_row += 1

    return start_row + 1


def add_homepage_stats_sheet(wb, homepage_stats, source_name, styles):
    ws = wb.create_sheet("홈페이지통계", 1)
    ws.merge_cells("A1:F1")
    ws["A1"] = "▶ 홈페이지 참여 현황 통계"
    ws["A1"].font = styles["title_font"]
    ws["A2"] = "홈페이지 /statistics/ 페이지에 표시되는 통계 항목과 동일한 기준입니다."
    ws["A3"] = f"데이터 기준 파일: {source_name}"

    widths = [28, 16, 16, 34, 16, 16]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    current_row = 5
    for key, label in CATEGORY_LABELS:
        data = homepage_stats[key]
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws.cell(current_row, 1, label)
        style_range(ws, current_row, 1, 6, fill=styles["category_fill"], font=styles["title_font"])
        current_row += 2

        summary_rows = [
            ["총 대상 세대", data["count"], "세대", f"단독: {data['single_count']} / 공동: {data['shared_count']}"],
            ["최종 동의 완료율", data["full_done_rate"], "%", f"{data['full_done_count']}세대 완료 / {data['remaining_household_count']}세대 남음"],
            ["일부 완료 세대", data["partial_done_count"], "세대", "공유자 추가 확인 시 전환 가능"],
            ["개인정보 동의 독려", data["privacy_needed_member_count"], "명", f"{data['privacy_needed_household_count']}세대 포함"],
        ]
        current_row = append_homepage_table(
            ws,
            "요약 지표",
            ["항목", "값", "단위", "보조 정보"],
            summary_rows,
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "세대 기준 최종 동의 상태",
            ["상태", "값"],
            [[item["label"], item["value"]] for item in data["household_stats"]],
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "소유 형태 비중",
            ["구분", "값"],
            [[item["label"], item["value"]] for item in data["owner_type_stats"]],
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "동의율 상승 우선순위",
            ["항목", "값", "단위", "보조 정보"],
            [[item["label"], item["value"], item.get("unit", ""), item.get("sub", "")] for item in data["followup_details"]],
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "소유자별 진행 상태",
            ["상태", "명"],
            [[item["label"], item["value"]] for item in data["participation_details"]],
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "최종 동의방식별 통계",
            ["최종 동의방식", "건수"],
            [[item["label"], item["value"]] for item in data["submission_details"]],
            current_row,
            styles,
        )

        dong_rows = [
            [item["label"], item["total"], item["done"], item["rate"]]
            for item in data["dong_stats"]
        ]
        current_row = append_homepage_table(
            ws,
            "동별 동의 현황 (전원 완료 세대 기준)",
            ["동", "전체 세대", "완료 세대", "동의율(%)"],
            dong_rows,
            current_row,
            styles,
        )

    return ws


def append_homepage_stats_sections(ws, homepage_stats, source_name, start_row, styles):
    current_row = start_row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    ws.cell(current_row, 1, "▶ 홈페이지 상세 통계")
    style_range(ws, current_row, 1, 8, font=styles["title_font"])
    current_row += 1
    ws.cell(current_row, 1, f"데이터 기준 파일: {source_name}")
    current_row += 2

    for key, label in CATEGORY_LABELS:
        data = homepage_stats[key]
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        ws.cell(current_row, 1, label)
        style_range(ws, current_row, 1, 8, fill=styles["category_fill"], font=styles["title_font"])
        current_row += 2

        summary_rows = [
            ["총 대상 세대", data["count"], "세대", f"단독: {data['single_count']} / 공동: {data['shared_count']}"],
            ["최종 동의 완료율", data["full_done_rate"], "%", f"{data['full_done_count']}세대 완료 / {data['remaining_household_count']}세대 남음"],
            ["일부 완료 세대", data["partial_done_count"], "세대", "공유자 추가 확인 시 전환 가능"],
            ["개인정보 동의 독려", data["privacy_needed_member_count"], "명", f"{data['privacy_needed_household_count']}세대 포함"],
        ]
        current_row = append_homepage_table(
            ws,
            "요약 지표",
            ["항목", "값", "단위", "보조 정보"],
            summary_rows,
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "세대 기준 최종 동의 상태",
            ["상태", "값"],
            [[item["label"], item["value"]] for item in data["household_stats"]],
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "소유 형태 비중",
            ["구분", "값"],
            [[item["label"], item["value"]] for item in data["owner_type_stats"]],
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "동의율 상승 우선순위",
            ["항목", "값", "단위", "보조 정보"],
            [[item["label"], item["value"], item.get("unit", ""), item.get("sub", "")] for item in data["followup_details"]],
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "소유자별 진행 상태",
            ["상태", "명"],
            [[item["label"], item["value"]] for item in data["participation_details"]],
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "최종 동의방식별 통계",
            ["최종 동의방식", "건수"],
            [[item["label"], item["value"]] for item in data["submission_details"]],
            current_row,
            styles,
        )

        current_row = append_homepage_table(
            ws,
            "동별 동의 현황 (전원 완료 세대 기준)",
            ["동", "전체 세대", "완료 세대", "동의율(%)"],
            [[item["label"], item["total"], item["done"], item["rate"]] for item in data["dong_stats"]],
            current_row,
            styles,
        )

    return current_row


def write_chart_source_table(ws, title, headers, rows, start_row, start_col, styles):
    ws.cell(start_row, start_col, title)
    ws.cell(start_row, start_col).font = styles["section_font"]
    start_row += 1

    for offset, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + offset, header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.border = styles["border"]
        cell.alignment = styles["center_align"]
    start_row += 1

    for row_values in rows:
        for offset, value in enumerate(row_values):
            cell = ws.cell(start_row, start_col + offset, value)
            cell.border = styles["border"]
        start_row += 1

    return start_row + 1


def add_homepage_charts(ws, homepage_stats, styles):
    total_stats = homepage_stats["total"]
    source_col = 27
    source_row = 1
    ws.column_dimensions[get_column_letter(source_col)].width = 28
    ws.column_dimensions[get_column_letter(source_col + 1)].width = 14

    household_rows = [[item["label"], item["value"]] for item in total_stats["household_stats"]]
    household_start = source_row + 1
    source_row = write_chart_source_table(
        ws,
        "그래프 데이터 - 세대 기준 최종 동의 상태",
        ["상태", "값"],
        household_rows,
        source_row,
        source_col,
        styles,
    )

    household_chart = PieChart()
    household_chart.title = "통합 세대 기준 최종 동의 상태"
    household_chart.add_data(
        Reference(ws, min_col=source_col + 1, min_row=household_start, max_row=household_start + len(household_rows)),
        titles_from_data=True,
    )
    household_chart.set_categories(
        Reference(ws, min_col=source_col, min_row=household_start + 1, max_row=household_start + len(household_rows))
    )
    household_chart.width = 13
    household_chart.height = 9
    ws.add_chart(household_chart, "Q55")

    owner_rows = [[item["label"], item["value"]] for item in total_stats["owner_type_stats"]]
    owner_start = source_row + 1
    source_row = write_chart_source_table(
        ws,
        "그래프 데이터 - 소유 형태 비중",
        ["구분", "값"],
        owner_rows,
        source_row,
        source_col,
        styles,
    )

    owner_chart = PieChart()
    owner_chart.title = "통합 소유 형태 비중"
    owner_chart.add_data(
        Reference(ws, min_col=source_col + 1, min_row=owner_start, max_row=owner_start + len(owner_rows)),
        titles_from_data=True,
    )
    owner_chart.set_categories(
        Reference(ws, min_col=source_col, min_row=owner_start + 1, max_row=owner_start + len(owner_rows))
    )
    owner_chart.width = 13
    owner_chart.height = 9
    ws.add_chart(owner_chart, "Y55")

    followup_rows = [[item["label"], item["value"]] for item in total_stats["followup_details"]]
    followup_start = source_row + 1
    source_row = write_chart_source_table(
        ws,
        "그래프 데이터 - 동의율 상승 우선순위",
        ["항목", "값"],
        followup_rows,
        source_row,
        source_col,
        styles,
    )

    followup_chart = BarChart()
    followup_chart.type = "bar"
    followup_chart.title = "통합 동의율 상승 우선순위"
    followup_chart.x_axis.title = "대상 수"
    followup_chart.y_axis.title = "항목"
    followup_chart.add_data(
        Reference(ws, min_col=source_col + 1, min_row=followup_start, max_row=followup_start + len(followup_rows)),
        titles_from_data=True,
    )
    followup_chart.set_categories(
        Reference(ws, min_col=source_col, min_row=followup_start + 1, max_row=followup_start + len(followup_rows))
    )
    followup_chart.width = 18
    followup_chart.height = 11
    ws.add_chart(followup_chart, "Q74")

    submission_rows = [[item["label"], item["value"]] for item in total_stats["submission_details"]]
    submission_start = source_row + 1
    write_chart_source_table(
        ws,
        "그래프 데이터 - 최종 동의방식별 통계",
        ["최종 동의방식", "건수"],
        submission_rows,
        source_row,
        source_col,
        styles,
    )

    submission_chart = BarChart()
    submission_chart.type = "bar"
    submission_chart.title = "통합 최종 동의방식별 통계"
    submission_chart.x_axis.title = "건수"
    submission_chart.y_axis.title = "최종 동의방식"
    submission_chart.add_data(
        Reference(ws, min_col=source_col + 1, min_row=submission_start, max_row=submission_start + len(submission_rows)),
        titles_from_data=True,
    )
    submission_chart.set_categories(
        Reference(ws, min_col=source_col, min_row=submission_start + 1, max_row=submission_start + len(submission_rows))
    )
    submission_chart.width = 13
    submission_chart.height = 9
    ws.add_chart(submission_chart, "Y74")


def build_dashboard_workbook(header, data_rows, source_name):
    column_map = build_stats_column_map(header, data_rows)
    padded_rows, final_submissions, household_summaries = build_dashboard_rows(data_rows, len(header), column_map)
    homepage_stats = build_homepage_stats(padded_rows, column_map, final_submissions)
    output_columns = resolve_output_columns(header, data_rows, column_map)
    output_header = [label for label, _, _ in output_columns]

    wb = openpyxl.Workbook()

    # --- 1. DATA_SOURCE SHEET ---
    ws_data = wb.active
    ws_data.title = "Data_Source"

    ws_data.append(output_header)

    for index, row_data in enumerate(padded_rows):
        ws_data.append(compact_output_row(row_data, output_columns, final_submissions[index]))

    title_font = Font(bold=True, size=16, color="002060")
    section_font = Font(bold=True, size=12, color="002060")
    header_font = Font(bold=True)
    tbl_header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    category_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    styles = {
        "title_font": title_font,
        "section_font": section_font,
        "header_font": header_font,
        "header_fill": tbl_header_fill,
        "category_fill": category_fill,
        "border": border,
        "center_align": center_align,
    }

    for cell in ws_data[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws_data.column_dimensions[get_column_letter(len(output_header))].width = 44
    for row in range(2, ws_data.max_row + 1):
        ws_data.cell(row=row, column=len(output_header)).alignment = Alignment(wrap_text=True, vertical="top")
    apply_auto_filter(ws_data, 1, end_col=len(output_header), end_row=ws_data.max_row)

    # --- 2. INTEGRATED STATS SHEET ---
    ws_summary = wb.create_sheet("통합통계", 0)
    ws_summary.merge_cells("A1:O1")
    ws_summary["A1"] = "▶ 통합 통계"
    ws_summary["A1"].font = title_font

    s_header = [
        "구분",
        "전체 인원",
        "인적 동의율",
        "전체 세대",
        "세대별 완료율",
        "전원 서면완료",
        "일부 동의",
        "단독 세대",
        "공동 세대",
        "전원 완료 세대",
        "미완료 세대",
        "개인정보 동의 독려(명)",
        "발송 후 미열람(명)",
        "열람 후 미완료(명)",
        "예약(명)",
    ]
    ws_summary.append(s_header)
    for header_cell in ws_summary[2]:
        header_cell.font = Font(bold=True)
        header_cell.fill = tbl_header_fill
        header_cell.border = border
        header_cell.alignment = center_align

    targets = [("통합", "", "total"), ("경남", "경남", "gyeongnam"), ("벽산", "벽산", "byeoksan")]
    for i, (label, apt, stats_key) in enumerate(targets):
        r = i + 3
        summary = summarize_complex(padded_rows, household_summaries, column_map, apt or None)
        homepage_data = homepage_stats[stats_key]

        ws_summary.cell(r, 1, label)
        ws_summary.cell(r, 2, summary["member_count"])
        ws_summary.cell(r, 3, summary["member_done_rate"]).number_format = "0.0%"
        ws_summary.cell(r, 4, summary["household_count"])
        ws_summary.cell(r, 5, summary["household_done_rate"]).number_format = "0.0%"
        ws_summary.cell(r, 6, summary["written_done_count"])
        ws_summary.cell(r, 7, summary["partial_count"])
        ws_summary.cell(r, 8, homepage_data["single_count"])
        ws_summary.cell(r, 9, homepage_data["shared_count"])
        ws_summary.cell(r, 10, homepage_data["full_done_count"])
        ws_summary.cell(r, 11, homepage_data["remaining_household_count"])
        ws_summary.cell(r, 12, homepage_data["privacy_needed_member_count"])
        ws_summary.cell(r, 13, homepage_data["sent_member_count"])
        ws_summary.cell(r, 14, homepage_data["viewed_member_count"])
        ws_summary.cell(r, 15, homepage_data["reserved_member_count"])
        for c in range(1, len(s_header) + 1):
            ws_summary.cell(r, c).border = border
            ws_summary.cell(r, c).alignment = Alignment(horizontal="center")

    ws_summary.append([])
    ws_summary.append(["▶ 동별 상세 지표 (서면 동의 관리)"])
    ws_summary.cell(ws_summary.max_row, 1).font = title_font

    d_header = ["동", "전체 세대", "전원 완료 세대", "세대 완료율", "서면 완료", "일부 동의", "50% 달성 부족분(서면)"]
    ws_summary.append(d_header)
    dong_header_row = ws_summary.max_row
    for header_cell in ws_summary[ws_summary.max_row]:
        header_cell.font = Font(bold=True)
        header_cell.fill = tbl_header_fill
        header_cell.border = border
        header_cell.alignment = center_align

    dong_summaries = summarize_dongs(household_summaries)
    for dong_summary in dong_summaries:
        d = dong_summary["dong"]
        r = ws_summary.max_row + 1
        ws_summary.cell(r, 1, display_dong_label(d))
        ws_summary.cell(r, 2, dong_summary["total"])
        ws_summary.cell(r, 3, dong_summary["full_done"])
        ws_summary.cell(r, 4, dong_summary["rate"]).number_format = "0.0%"
        ws_summary.cell(r, 5, dong_summary["written_done"])
        ws_summary.cell(r, 6, dong_summary["partial"])
        ws_summary.cell(r, 7, dong_summary["shortage"])
        for c in range(1, 8):
            ws_summary.cell(r, c).border = border
    dong_last_row = ws_summary.max_row
    apply_auto_filter(ws_summary, dong_header_row, end_col=len(d_header), end_row=dong_last_row)

    widths = [14, 12, 12, 12, 14, 14, 12, 12, 12, 14, 12, 18, 18, 18, 12]
    for i, width in enumerate(widths, start=1):
        ws_summary.column_dimensions[get_column_letter(i)].width = width
    add_summary_charts(ws_summary, dong_header_row, dong_last_row)
    append_homepage_stats_sections(ws_summary, homepage_stats, source_name, dong_last_row + 3, styles)
    add_homepage_charts(ws_summary, homepage_stats, styles)

    # --- 3. DONG SHEETS ---
    dong_groups = defaultdict(list)
    for index, row in enumerate(padded_rows):
        dong_groups[cell_value(row, column_map, "dong") or "기타"].append(index)

    for dong_summary in dong_summaries:
        d = dong_summary["dong"]
        ws = wb.create_sheet(title=display_dong_label(d)[:31])
        ws.append(output_header)
        for row_index in dong_groups[d]:
            row_data = padded_rows[row_index]
            ws.append(compact_output_row(row_data, output_columns, final_submissions[row_index]))
        memo_column = len(output_header)
        ws.column_dimensions[get_column_letter(memo_column)].width = 44
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=memo_column).alignment = Alignment(wrap_text=True, vertical="top")
        apply_auto_filter(ws, 1, end_col=len(output_header), end_row=ws.max_row)

    return wb


def save_dashboard_workbook(header, data_rows, source_name, output_path):
    wb = build_dashboard_workbook(header, data_rows, source_name)
    wb.save(output_path)
    return wb


def main():
    input_path = resolve_source_path(PROJECT_ROOT)
    output_path = PROJECT_ROOT / 'raw_data' / '발송현황_실시간대시보드_최종.xlsx'

    if not os.path.exists(input_path):
        print(f"Error: {input_path} not found.")
        return

    try:
        rows = load_xlsx_rows(input_path)
        if not rows: return
        header = rows[0]
        data_rows = rows[1:]
        save_dashboard_workbook(header, data_rows, input_path.name, output_path)
        print(f"ULTRA-STABLE Dashboard Created: {output_path}")

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
